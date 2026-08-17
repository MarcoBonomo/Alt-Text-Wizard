import streamlit as st
from anthropic import Anthropic
from PIL import Image
import base64
from io import BytesIO
import pandas as pd
import time
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows

# Set page config
st.set_page_config(
    page_title="Alt-Text Wizard",
    page_icon="🧙‍♂️",
    layout="wide"
)

# Initialize session state
if 'total_cost' not in st.session_state:
    st.session_state.total_cost = 0.0
if 'total_tokens' not in st.session_state:
    st.session_state.total_tokens = {'input': 0, 'output': 0}
if 'api_calls' not in st.session_state:
    st.session_state.api_calls = 0
if 'total_execution_time' not in st.session_state:
    st.session_state.total_execution_time = 0.0
if 'current_result' not in st.session_state:
    st.session_state.current_result = None
if 'bulk_results' not in st.session_state:
    st.session_state.bulk_results = None
if 'selected_model' not in st.session_state:
    st.session_state.selected_model = "claude-opus-4-20250514"

# Model constants - Opus first, then Haiku
MODELS = {
    "claude-opus-4-20250514": "Claude 4 Opus (Premium Quality)",
    "claude-3-5-haiku-20241022": "Claude 3.5 Haiku (Fast & Cheap)"
}

# Pricing constants
PRICING = {
    "claude-opus-4-20250514": {
        "input": 15.00 / 1_000_000,
        "output": 75.00 / 1_000_000
    },
    "claude-3-5-haiku-20241022": {
        "input": 0.80 / 1_000_000,
        "output": 4.00 / 1_000_000
    }
}

# BATCHING CONFIGURATION
MAX_IMAGES_PER_BATCH = 5  # Adjust this based on your needs (3-10 recommended)

# Banned words lists
BANNED_VERBS = [
    'love', 'adore', 'relish', "don't miss", 'dont miss'
]

BANNED_ADJECTIVES = [
    'luxurious', 'best', 'better', 'worse', 'worst',
    'safe', 'safer', 'safest', 'fastest',
    'fresh', 'seamless', 'lovely', 'adorable',
    'sophisticated', 'beautiful', 'satisfying'
]

BANNED_PRICE_TERMS = [
    'cheap', 'affordable', 'low cost', 'low-cost',
    'lower price', 'lowest prices', 'lowest price',
    'budget-friendly', 'budget friendly', 'economical', 'inexpensive'
]

# Combine all banned words
ALL_BANNED_WORDS = BANNED_VERBS + BANNED_ADJECTIVES + BANNED_PRICE_TERMS

# Multi-Product Override Logic - Number to word conversion
NUMBER_TO_WORD = {
    1: "One",
    2: "Two",
    3: "Three",
    4: "Four",
    5: "Five",
    6: "Six",
    7: "Seven",
    8: "Eight",
    9: "Nine",
    10: "Ten"
}

# ============================================================================
# REQUIREMENT B: PMI Product Identification Guide
# ============================================================================
PMI_PRODUCT_GUIDE = """# PHILIP MORRIS INTERNATIONAL (PMI) SMOKE-FREE PRODUCT IDENTIFICATION GUIDE

## IQOS ILUMA DEVICES (Heated Tobacco - Bladeless Induction)

### Device Models:
- **IQOS ILUMA / ILUMA PRIME**: Two-piece system with separate holder and pocket charger. Premium aluminum body with bladeless SMARTCORE INDUCTION SYSTEM. Uses TEREA or LEVIA sticks ONLY.
  
- **IQOS ILUMA ONE**: All-in-one compact single-unit design. 20 uses per charge. Uses TEREA or LEVIA sticks ONLY.

- **IQOS ILUMA i / i PRIME / i ONE**: Latest generation featuring TOUCH SCREEN on holder, Pause Mode, and FlexPuff capabilities. Uses TEREA or LEVIA sticks ONLY.

**Visual identifiers**: Bladeless design, SMARTCORE INDUCTION SYSTEM, modern sleek metallic body, premium finish.

---

## OLDER IQOS DEVICES (Blade-based Heating)

- **IQOS 2.4, IQOS 3, IQOS 3 DUO**: Previous generation devices with internal heating BLADE. Uses HEETS sticks ONLY. NOT compatible with TEREA/LEVIA.

**Visual identifiers**: Visible blade slot, older design language, HEETS compatibility only.

---

## TOBACCO STICKS & CONSUMABLES

### TEREA Sticks (ILUMA-exclusive)
- Small cylindrical sticks in packs of 20
- Contains internal metal heating element
- **ONLY compatible with IQOS ILUMA devices**
- Modern vibrant packaging with TEREA branding
- **Color coding by flavor**:
  - Amber = orange packaging
  - Yellow = gold packaging
  - Russet = purple packaging
  - Blue = dark blue packaging (menthol)
  - Green = bright green packaging (menthol)

### LEVIA Sticks (ILUMA-exclusive, Tobacco-free)
- **TOBACCO-FREE nicotine sticks** with similar form factor to TEREA
- **ONLY compatible with IQOS ILUMA devices**
- Uses same SMARTCORE induction technology
- Cleaner/lighter packaging design compared to TEREA
- Prominent LEVIA branding
- Positioned as tobacco-free alternative

### HEETS Sticks (Legacy devices only)
- Tobacco sticks for **OLDER IQOS devices** (2.4, 3, 3 DUO) and lil SOLID
- **NOT compatible with IQOS ILUMA**
- HEETS branding (called "HeatSticks" in Japan)
- Works with blade-heating technology
- **Color coding**: Amber, Yellow, Turquoise, Blue, Bronze, Sienna, etc.

### Fiit Sticks
- Tobacco sticks designed for **lil SOLID devices**
- Similar form factor to HEETS
- **NOT compatible with IQOS ILUMA**

---

## VEEV PRODUCTS (Vaping/E-cigarettes)

### Devices:
- **VEEV ONE**: Compact aluminum pod-based vape. Sleek rectangular body, magnetic pod attachment, USB-C charging, side LED indicators.

- **VEEV Ultra**: Premium/advanced pod-based vape. Larger form factor than VEEV ONE.

- **VEEV NOW**: Disposable single-use vape. Pen-like shape, puff-activated, no charging required.

### Consumables:
- **VEEV ONE pods**: Small prefilled e-liquid pods with colored ring indicating flavor. Magnetic attachment.
- **VEEV Ultra pods**: Pods specifically for VEEV Ultra device.

**Visual identifiers**: Aluminum body, closed pod system, VEEV branding, slim modern form factor, LED indicators.

---

## lil SOLID DEVICES (Budget Heated Tobacco)

- **lil SOLID / SOLID 2.0 / SOLID Ez**: All-in-one compact devices with single-button operation
- Ceramic PIN heating technology (not blade, not induction)
- Uses **Fiit sticks or HEETS** (NOT TEREA or LEVIA)
- Budget-friendly alternative to IQOS

**Visual identifiers**: Compact all-in-one design, single button, simpler aesthetic.

---

## ZYN NICOTINE POUCHES (Oral Nicotine)

- Small round or rectangular plastic container/can with click-open lid
- Contains small pillow-shaped nicotine pouches
- Strength indicated by dots/numbers on packaging
- Prominent ZYN logo on container
- Tobacco-free oral nicotine delivery

**Visual identifiers**: Cylindrical can, snap lid, ZYN branding, strength indicators.

---

## MANDATORY SPELLING & NAMING CONVENTIONS

**Use EXACT spelling - these are brand names:**

| ✅ CORRECT | ❌ INCORRECT |
|-----------|-------------|
| IQOS | Ikos, iQOS, Iqos |
| ILUMA | Illuma, Iluma |
| ILUMA i | ILUMA I, Iluma i (lowercase 'i') |
| TEREA | Tarea, Teria, Terrea |
| LEVIA | Levya, Livia, Levia |
| HEETS | Heats, HEETs, Heets |
| VEEV ONE | Veev 1, VeevOne, Veev one |
| VEEV Ultra | Veev ultra, VEEV ULTRA, VeevUltra |
| lil SOLID | Lil Solid, LIL SOLID, lilSOLID (lowercase 'lil', uppercase 'SOLID') |
| ZYN | Zyn, zyn, ZYn (all caps) |
| Fiit | Fitt, FIIT, fit |

---

## COMPATIBILITY MATRIX

| Device | Compatible Sticks | ❌ NOT Compatible |
|--------|------------------|------------------|
| IQOS ILUMA (all variants) | TEREA, LEVIA | HEETS, Fiit |
| IQOS 2.4/3/3 DUO | HEETS | TEREA, LEVIA, Fiit |
| lil SOLID (all variants) | Fiit, HEETS | TEREA, LEVIA |
| VEEV ONE | VEEV ONE pods | VEEV Ultra pods |
| VEEV Ultra | VEEV Ultra pods | VEEV ONE pods |

---

## ALT-TEXT GENERATION INSTRUCTIONS

When generating alt-text:

1. **Identify the specific product** using visual cues from the guide above
2. **Use exact product names** per the spelling rules
3. **Include device generation/model** when identifiable (e.g., "IQOS ILUMA i" not just "IQOS")
4. **Specify stick type** for consumables and mention color/flavor when visible
5. **Note compatibility** if relevant (e.g., "TEREA sticks for IQOS ILUMA devices")
6. **Describe visual elements**: color, packaging design, key features
7. **Be concise but comprehensive** - typically 100-150 characters for product images

### Examples:
- ✅ "IQOS ILUMA PRIME holder in moonlight silver with pocket charger, bladeless heated tobacco device"
- ✅ "TEREA Amber tobacco sticks pack (orange packaging) for IQOS ILUMA devices, 20 sticks"
- ✅ "VEEV ONE compact pod vape in black with USB-C charging, side LED indicators"
- ❌ "Ikos device" (incorrect spelling)
- ❌ "TEREA sticks for IQOS" (incomplete - should specify ILUMA)
"""

# Enhanced alt-text generation prompt with PMI product guide
ALT_TEXT_PROMPT = f"""You are an expert in visual accessibility and e-commerce product description. Generate accurate, vivid, and highly detailed alt-text for this product image.

**PRODUCT IDENTIFICATION REFERENCE:**

{PMI_PRODUCT_GUIDE}

**CRITICAL REQUIREMENTS:**

1. **Start with "A" or "An" OR Number Word**
   - Single product: Start with "A" or "An"
     • Example: "A sleek black smartphone..."
   - Multiple products (2+): Start with NUMBER WORD, NOT "A" or "An"
     • Example: "Two smartphones in silver and black"
     • Example: "Three devices with LED indicators"
     • WRONG: "A two smartphones..." or "An three devices..."

2. **Human Presence - Combined Description When Holding/Using Device**
   - IF person is holding/using a device: Combine in ONE sentence with "holding" or "using"
     • Format: "A person wearing [clothing] holding a [device description]"
     • Example: "A person wearing a yellow shirt holding a black and silver device with metallic finish"
     • Example: "A person in blue jeans using a smartphone with edge-to-edge display"
   - IF person is NOT holding/using (just visible): Use separate sentences
     • Format: "A person wearing [clothing]. A [product description]."
     • Example: "A person wearing a gray shirt. A water bottle on a table."
   - NEVER use format like: "A person in yellow shirt. A device in hand." — This is WRONG
   - ALWAYS combine when person holds/uses device: "A person in yellow shirt holding a device"

3. **Multi-Product Detection**
   - If multiple devices/products are visible (2+), specify the count
   - Start with NUMBER as WORD (Two, Three, Four...), NOT "A" or "An"
   - Use format: "[Number as word] [product type]s"
   - Examples:
     • "Two smartphones in silver and black finish"
     • "Three vaping devices with LED indicators"
     • "Four bottles arranged in a row"

4. **Enhanced Nuanced Details**
   - **Textures**: matte, glossy, brushed, textured, smooth, ribbed, pebbled, hammered
   - **Finishes**: polished, anodized, powder-coated, chrome-plated, satin, mirror-finish
   - **Proportions**: slim, compact, oversized, petite, elongated, chunky
   - **Surface details**: embossed logo, debossed text, laser-etched pattern, perforated panels
   - **Hardware specifics**: rose gold zipper pull, gunmetal buckle, silver-tone D-rings
   - **Material qualities**: full-grain leather, recycled polyester, tempered glass, cast aluminum
   - **Color depth**: deep navy (not just "blue"), charcoal gray (not just "gray"), rose gold (not just "gold")
   - **Design elements**: contrast stitching, quilted panels, mesh inserts, ventilation holes
   - **Brand marks**: embossed logo on clasp, debossed brand name on side, etched serial number
   - **Functional features**: magnetic closure, snap button, adjustable strap, removable component

5. **Accessibility-First Language**
   - NO subjective claims or banned words
   - NO marketing language: "best-selling," "must-have," "perfect for"
   - NO calls to action: "shop now," "buy today," "get yours"
   - Use OBJECTIVE, DESCRIPTIVE language only
   - **STRICTLY AVOID THESE BANNED WORDS**: love, adore, relish, don't miss, luxurious, best, better, worse, worst, safe, safer, safest, fastest, fresh, seamless, lovely, adorable, sophisticated, beautiful, satisfying, cheap, affordable, low cost, lower price, lowest prices, budget-friendly, economical, inexpensive

6. **Length Limit**
   - **MAXIMUM 150 characters total**
   - Prioritize the most visually distinctive features
   - Be concise but descriptive

7. **Brand Name Handling**
   - **NEVER repeat brand names** (e.g., avoid "VEEV VEEV" or "Veev veev")
   - Use brand name only ONCE
   - If brand appears multiple times, use only the properly capitalized version
   - **Follow PMI spelling conventions exactly** (IQOS, ILUMA, TEREA, LEVIA, VEEV, etc.)

8. **Contextual Props**
   - ONLY mention props if they provide scale, demonstrate use, or are integral to understanding
   - Otherwise, focus entirely on the product

**STRUCTURE TEMPLATES:**

Single product WITHOUT person:
"A [material] [product type] in [specific color], featuring [distinctive design elements]"

Single product WITH person holding/using it (COMBINED):
"A person wearing [clothing] holding a [material] [product type] in [specific color]"

Single product WITH person NOT holding it (SEPARATED):
"A person wearing [clothing]. A [material] [product type] on [location]."

Multiple products WITHOUT person (START WITH NUMBER WORD):
"[Number as word] [product type]s in [colors/finishes], featuring [distinctive elements]"

Multiple products WITH person:
"A person wearing [clothing] holding [number as word] [product type]s in [colors/finishes]"

**EXAMPLES:**

✅ CORRECT (single product, no person):
"A stainless steel water bottle in matte black with powder-coated exterior and embossed logo"

✅ CORRECT (person holding device - COMBINED):
"A person wearing a yellow shirt holding a black and silver device with metallic finish"

✅ CORRECT (person holding device - COMBINED):
"A person in blue jeans holding a smartphone with edge-to-edge display and silver frame"

✅ CORRECT (multiple products, no person - STARTS WITH NUMBER):
"Two smartphones in silver and black with edge-to-edge displays"

✅ CORRECT (multiple products with person):
"A person wearing a blue hoodie holding three devices in metallic finishes"

✅ CORRECT (PMI product):
"IQOS ILUMA PRIME holder in moonlight silver with pocket charger, bladeless heated tobacco device"

✅ CORRECT (PMI product):
"TEREA Amber tobacco sticks pack in orange packaging for IQOS ILUMA devices"

❌ WRONG (person and device separated when person is holding it):
"A person wearing a yellow shirt. A black and silver device in hand."

❌ WRONG (multiple products starting with "A"):
"A two smartphones in silver and black"

❌ WRONG (banned words):
"A luxurious premium water bottle that's perfect for your lifestyle"

❌ WRONG (too long, over 150 characters):
"A stainless steel insulated water bottle in matte black finish, featuring a powder-coated exterior, knurled grip texture on the lower half, screw-top lid with silicone seal, and an embossed mountain logo on the front panel"

❌ WRONG (brand repetition):
"A VEEV VEEV device in silver finish"

❌ WRONG (incorrect PMI spelling):
"An Ikos Illuma device in silver" (should be "IQOS ILUMA")

**OUTPUT FORMAT:**
Return ONLY the alt-text string (max 150 characters)—no labels, no markdown, no explanations, no quotation marks."""

# Sidebar for API key
with st.sidebar:
    st.header("⚙️ Configuration")
    api_key = st.text_input(
        "Anthropic API Key",
        type="password",
        help="Enter your Anthropic API key. It will only be stored for this session."
    )
    
    if api_key:
        st.success("API Key provided ✓")
    else:
        st.warning("Please enter your API Key to continue")
    
    st.markdown("---")
    
    # Model selection
    st.header("🤖 Model Selection")
    
    model_choice = st.radio(
        "Choose Model:",
        options=list(MODELS.keys()),
        format_func=lambda x: MODELS[x],
        index=0,
        help="**Opus 4**: Premium quality, best visual understanding (default)\n\n**Haiku 3.5**: Fast and cost-effective alternative"
    )
    
    st.session_state.selected_model = model_choice
    
    # Model info based on selection
    if "opus" in model_choice:
        st.info("✨ **Premium Quality (Default)**\nClaude 4 Opus - Best visual understanding for critical tasks")
    else:
        st.info("⚡ **Cost-Effective Alternative**\nClaude 3.5 Haiku - Fast, accurate, and budget-friendly")
    
    # Show pricing for selected model
    with st.expander("💰 Pricing Details"):
        st.write(f"**{MODELS[model_choice]}**")
        st.write(f"Input: ${PRICING[model_choice]['input'] * 1_000_000:.2f} / 1M tokens")
        st.write(f"Output: ${PRICING[model_choice]['output'] * 1_000_000:.2f} / 1M tokens")
        
        # Show comparison
        st.markdown("---")
        st.caption("**Model Comparison:**")
        for model_id, model_name in MODELS.items():
            input_price = PRICING[model_id]['input'] * 1_000_000
            output_price = PRICING[model_id]['output'] * 1_000_000
            st.caption(f"{model_name}: ${input_price:.2f}/${output_price:.2f}")
        
        st.markdown("---")
        st.caption("**💡 Tip:** Opus 4 is default for quality. Switch to Haiku 3.5 for faster processing.")
    
    st.markdown("---")
    
    # Batch size configuration
    st.header("⚙️ Batch Settings")
    
    batch_size = st.slider(
        "Images per API call:",
        min_value=3,
        max_value=10,
        value=MAX_IMAGES_PER_BATCH,
        help="Adjust based on image size. Use 3-5 for large images, 5-10 for smaller images."
    )
    MAX_IMAGES_PER_BATCH = batch_size
    
    st.caption(f"💡 Current: {MAX_IMAGES_PER_BATCH} images/batch")
    
    st.markdown("---")
    
    # Cost tracking section
    st.header("📊 Session Statistics")
    
    col1, col2 = st.columns(2)
    with col1:
        st.metric("API Calls", st.session_state.api_calls)
    with col2:
        st.metric("⏱️ Total Time", f"{st.session_state.total_execution_time:.2f}s")
    
    st.metric("💰 Total Cost", f"${st.session_state.total_cost:.4f}")
    
    # Average metrics
    if st.session_state.api_calls > 0:
        avg_time = st.session_state.total_execution_time / st.session_state.api_calls
        avg_cost = st.session_state.total_cost / st.session_state.api_calls
        
        col3, col4 = st.columns(2)
        with col3:
            st.metric("Avg Time", f"{avg_time:.2f}s")
        with col4:
            st.metric("Avg Cost", f"${avg_cost:.4f}")
    
    with st.expander("🔢 Token Details"):
        st.write(f"**Input tokens:** {st.session_state.total_tokens['input']:,}")
        st.write(f"**Output tokens:** {st.session_state.total_tokens['output']:,}")
        st.write(f"**Total tokens:** {st.session_state.total_tokens['input'] + st.session_state.total_tokens['output']:,}")
    
    if st.button("Reset Statistics", use_container_width=True):
        st.session_state.total_cost = 0.0
        st.session_state.total_tokens = {'input': 0, 'output': 0}
        st.session_state.api_calls = 0
        st.session_state.total_execution_time = 0.0
        st.session_state.current_result = None
        st.session_state.bulk_results = None
        st.rerun()
    
    st.markdown("---")
    st.markdown("### About")
    st.markdown("""
    This app uses Anthropic's Claude Vision models to generate detailed, accessible alt text for product images.
    
    **Features:**
    - Single image processing
    - **Smart batch processing** (configurable batch size)
    - **PMI product identification** (IQOS, TEREA, VEEV, etc.)
    - Two model options (Opus/Haiku)
    - Accessibility-focused descriptions
    - Enhanced detail recognition
    - Person detection & combined descriptions
    - Multi-product detection
    - Download results as CSV/Excel
    - Excel export with thumbnails
    - Cost estimation & tracking
    - Banned word filtering
    - 150 character limit
    """)

# Title and description
st.title("Alt-Text Wizard 🧙‍♂️✨")
st.markdown("**Get instant, accessible alt text for your images—powered by Anthropic's Claude Vision API and expert accessibility guidelines.**")

# Function to check for banned words
def check_banned_words(text):
    """Check if text contains any banned words and return list of found banned words"""
    if not text:
        return []
    
    text_lower = text.lower()
    found_banned = []
    
    for word in ALL_BANNED_WORDS:
        # Use word boundaries for accurate matching
        pattern = r'\b' + re.escape(word.lower()) + r'\b'
        if re.search(pattern, text_lower):
            found_banned.append(word)
    
    return found_banned

# Function to encode image to base64
def encode_image(image_file):
    """Convert uploaded file to base64 string"""
    return base64.b64encode(image_file.read()).decode('utf-8')

# Function to get image media type
def get_image_media_type(filename):
    """Determine media type from filename"""
    ext = filename.lower().split('.')[-1]
    media_types = {
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'webp': 'image/webp'
    }
    return media_types.get(ext, 'image/jpeg')

# Function to create thumbnails
def create_thumbnail(image_file, size=(80, 80)):
    """Create a thumbnail from uploaded file"""
    image_file.seek(0)
    img = Image.open(image_file)
    img.thumbnail(size, Image.Resampling.LANCZOS)
    return img

# Function to create Excel with images
def create_excel_with_images(results, uploaded_files):
    """Create Excel file with embedded image thumbnails"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Alt Text Results"
    
    # Headers
    headers = ["Thumbnail", "Filename", "Alt Text", "Character Count", "Est. Cost ($)", "Est. Tokens"]
    ws.append(headers)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Thumbnail
    ws.column_dimensions['B'].width = 30  # Filename
    ws.column_dimensions['C'].width = 80  # Alt Text
    ws.column_dimensions['D'].width = 15  # Character Count
    ws.column_dimensions['E'].width = 15  # Est. Cost
    ws.column_dimensions['F'].width = 15  # Est. Tokens
    
    # Add data rows with images
    for idx, (result, file) in enumerate(zip(results, uploaded_files), start=2):
        # Create thumbnail
        file.seek(0)
        img = Image.open(file)
        img.thumbnail((80, 80), Image.Resampling.LANCZOS)
        
        # Save thumbnail to BytesIO
        img_byte_arr = BytesIO()
        img.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)
        
        # Create openpyxl image
        openpyxl_img = OpenpyxlImage(img_byte_arr)
        openpyxl_img.width = 80
        openpyxl_img.height = 80
        
        # Add image to cell
        ws.add_image(openpyxl_img, f'A{idx}')
        
        # Set row height to fit image
        ws.row_dimensions[idx].height = 60
        
        # Add text data
        ws[f'B{idx}'] = result['Filename']
        ws[f'C{idx}'] = result['Alt Text']
        ws[f'D{idx}'] = result['Character Count']
        ws[f'E{idx}'] = result['Est. Cost ($)']
        ws[f'F{idx}'] = result['Est. Tokens']
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

# Function to calculate cost
def calculate_cost(usage, model):
    """Calculate cost based on token usage"""
    input_cost = usage.input_tokens * PRICING[model]["input"]
    output_cost = usage.output_tokens * PRICING[model]["output"]
    total_cost = input_cost + output_cost
    
    return {
        "input_tokens": usage.input_tokens,
        "output_tokens": usage.output_tokens,
        "total_tokens": usage.input_tokens + usage.output_tokens,
        "input_cost": input_cost,
        "output_cost": output_cost,
        "total_cost": total_cost
    }

# Enhanced article prefix logic
def ensure_article_prefix(alt_text):
    """
    Ensure alt text starts correctly:
    - Single product: "A" or "An"
    - Multiple products (2+): Number word (Two, Three, etc.)
    """
    alt_text = alt_text.strip()
    
    # Check if it starts with a number word (Two, Three, Four, etc.)
    number_word_pattern = r'^(Two|Three|Four|Five|Six|Seven|Eight|Nine|Ten)\s'
    if re.match(number_word_pattern, alt_text, re.IGNORECASE):
        # Multiple products - capitalize number word properly
        match = re.match(number_word_pattern, alt_text, re.IGNORECASE)
        if match:
            number_word = match.group(1)
            rest = alt_text[len(number_word):].strip()
            return f"{number_word.capitalize()} {rest}"
        return alt_text
    
    # Check if it already starts with A/An (single product)
    if re.match(r'^(A|An)\s', alt_text, re.IGNORECASE):
        # Capitalize properly
        if alt_text.startswith('a '):
            return 'A' + alt_text[1:]
        elif alt_text.startswith('an '):
            return 'An' + alt_text[2:]
        return alt_text
    
    # If doesn't start with number word or A/An, add appropriate article for single product
    # If starts with vowel sound, use "An"
    if re.match(r'^[aeiouAEIOU]', alt_text):
        return f"An {alt_text}"
    
    # Otherwise use "A"
    return f"A {alt_text}"

# Function to remove brand name repetition
def remove_brand_repetition(text):
    """Remove repeated brand names (e.g., 'VEEV VEEV' -> 'VEEV')"""
    words = text.split()
    cleaned_words = []
    i = 0
    
    while i < len(words):
        # Check if current word matches next word (case-insensitive)
        if i < len(words) - 1 and words[i].upper() == words[i + 1].upper():
            # Keep the properly capitalized version (prefer all caps or title case)
            if words[i].isupper():
                cleaned_words.append(words[i])
            elif words[i + 1].isupper():
                cleaned_words.append(words[i + 1])
            elif words[i][0].isupper():
                cleaned_words.append(words[i])
            else:
                cleaned_words.append(words[i + 1])
            i += 2  # Skip both words
        else:
            cleaned_words.append(words[i])
            i += 1
    
    return ' '.join(cleaned_words)

# Multi-Product Detection and Override Logic
def detect_device_count(alt_text):
    """Detect number of devices mentioned in alt text"""
    # Look for patterns like "two devices", "three smartphones", etc.
    number_pattern = r'\b(two|three|four|five|six|seven|eight|nine|ten|\d+)\s+(device|smartphone|bottle|watch|product|item)'
    match = re.search(number_pattern, alt_text.lower())
    
    if match:
        number_str = match.group(1)
        # Convert word to number
        word_to_num = {v.lower(): k for k, v in NUMBER_TO_WORD.items()}
        if number_str in word_to_num:
            return word_to_num[number_str]
        elif number_str.isdigit():
            return int(number_str)
    
    # Check for plural indicators
    plural_pattern = r'\b(devices|smartphones|bottles|watches|products|items)\b'
    if re.search(plural_pattern, alt_text.lower()):
        # Default to 2 if plural but no specific number found
        return 2
    
    return 1  # Default to single device

def apply_multi_product_override(alt_text, device_name, device_count):
    """
    Apply device name override with proper pluralization for multiple products
    """
    if not device_name or not device_name.strip():
        return alt_text
    
    device_name = device_name.strip()
    
    # Patterns to replace with device name
    if device_count > 1:
        # Multiple devices - use number + plural form
        number_word = NUMBER_TO_WORD.get(device_count, str(device_count))
        device_plural = f"{device_name}s" if not device_name.endswith('s') else device_name
        replacement = f"{number_word} {device_plural}"
        
        # Replace patterns for multiple devices
        patterns = [
            (r'\b(two|three|four|five|six|seven|eight|nine|ten|\d+)\s+(device|smartphone|bottle|watch|product|item)s?\b', replacement),
            (r'\b(device|smartphone|bottle|watch|product|item)s\b', device_plural),
            (r'\bmultiple\s+(device|smartphone|bottle|watch|product|item)s?\b', replacement),
        ]
    else:
        # Single device
        replacement = device_name
        patterns = [
            (r'\b(vape\s+pen|vaping\s+device|vaporizer|e-cigarette|electronic\s+cigarette)\b', replacement),
            (r'\b(device|smartphone|bottle|watch|product|item)\b', replacement),
            (r'\b([A-Z]{2,}\s+[A-Z]{2,}(?:\s+[A-Z0-9]+)?)\b', replacement),
            (r'\b([A-Z]{3,})\b', replacement),
        ]
    
    for pattern, repl in patterns:
        new_text = re.sub(pattern, repl, alt_text, count=1, flags=re.IGNORECASE)
        if new_text != alt_text:
            new_text = re.sub(r'\s+', ' ', new_text).strip()
            new_text = remove_brand_repetition(new_text)
            return new_text
    
    return alt_text

# Function to substitute device name in alt text
def substitute_device_name(alt_text, device_name):
    """Replace generic product terms or redundant device names with specific device name."""
    if not device_name or not device_name.strip():
        return alt_text
    
    device_name = device_name.strip()
    
    # First remove any brand repetition in the alt text
    alt_text = remove_brand_repetition(alt_text)
    
    # Detect device count and apply multi-product override
    device_count = detect_device_count(alt_text)
    alt_text = apply_multi_product_override(alt_text, device_name, device_count)
    
    # Remove any remaining brand repetition after substitution
    alt_text = remove_brand_repetition(alt_text)
    
    return alt_text

# Function to enforce 150 character limit
def enforce_character_limit(alt_text, max_length=150):
    """Truncate alt text to max_length characters at last complete word"""
    if len(alt_text) <= max_length:
        return alt_text
    
    # Truncate at last space before max_length
    truncated = alt_text[:max_length]
    last_space = truncated.rfind(' ')
    
    if last_space > 0:
        return truncated[:last_space].strip()
    
    # If no space found, hard truncate
    return truncated.strip()

# Enhanced person description logic
def combine_person_holding_device(alt_text):
    """
    Combine person and device descriptions when person is holding/using device.
    """
    # Pattern 1: Detect separated person and device descriptions where person is holding
    pattern1 = r'(A\s+person\s+(?:wearing\s+)?[^.]+)\.\s*(A\s+[^.]+(?:in\s+hand|in\s+their\s+hand|being\s+held)[^.]*)'
    match1 = re.search(pattern1, alt_text, re.IGNORECASE)
    
    if match1:
        person_part = match1.group(1).strip()
        device_part = match1.group(2).strip()
        
        # Remove "in hand", "in their hand", "being held" from device part
        device_part = re.sub(r'\s+in\s+hand.*$', '', device_part, flags=re.IGNORECASE)
        device_part = re.sub(r'\s+in\s+their\s+hand.*$', '', device_part, flags=re.IGNORECASE)
        device_part = re.sub(r'\s+being\s+held.*$', '', device_part, flags=re.IGNORECASE)
        device_part = device_part.strip()
        
        # Change "A device" to "a device" (lowercase) for combination
        device_part_lower = device_part[0].lower() + device_part[1:] if device_part else device_part
        
        # Combine with "holding"
        combined = f"{person_part} holding {device_part_lower}"
        
        # Get any remaining text after the device description
        rest = alt_text.replace(match1.group(0), '').strip()
        
        if rest:
            return f"{combined}. {rest}"
        return combined
    
    # Pattern 2: Look for person description followed by device with possessive indicators
    pattern2 = r'(A\s+person\s+(?:wearing\s+)?[^.]+)\.\s*(?:Their|His|Her)\s+([^.]+)'
    match2 = re.search(pattern2, alt_text, re.IGNORECASE)
    
    if match2:
        person_part = match2.group(1).strip()
        device_part = match2.group(2).strip()
        
        # Combine with "holding"
        combined = f"{person_part} holding a {device_part}"
        
        # Get any remaining text
        rest = alt_text.replace(match2.group(0), '').strip()
        
        if rest:
            return f"{combined}. {rest}"
        return combined
    
    # Pattern 3: Generic pattern for person + device in separate sentences
    pattern3 = r'(A\s+person\s+(?:wearing|in)\s+[^.]+)\.\s*(A\s+[^.]+device[^.]*)'
    match3 = re.search(pattern3, alt_text, re.IGNORECASE)
    
    if match3:
        person_part = match3.group(1).strip()
        device_part = match3.group(2).strip()
        
        # Change "A device" to "a device" (lowercase)
        device_part_lower = device_part[0].lower() + device_part[1:] if device_part else device_part
        
        # Combine with "holding"
        combined = f"{person_part} holding {device_part_lower}"
        
        # Get any remaining text
        rest = alt_text.replace(match3.group(0), '').strip()
        
        if rest:
            return f"{combined}. {rest}"
        return combined
    
    return alt_text

# Enhanced Person Description Separation
def separate_person_description(alt_text):
    """
    Ensure person descriptions are properly formatted:
    - If person holding/using device: Combine in one sentence
    - If person NOT holding: Keep separate sentences
    """
    # First, try to combine if person is holding device
    alt_text = combine_person_holding_device(alt_text)
    
    # Check if person is mentioned and NOT already combined with "holding" or "using"
    person_pattern = r'(A\s+person\s+(?:wearing|in)\s+[^.]+)'
    has_person = re.search(person_pattern, alt_text, re.IGNORECASE)
    has_holding = re.search(r'\bholding\b|\busing\b', alt_text, re.IGNORECASE)
    
    if has_person and not has_holding:
        # Person is NOT holding - ensure proper separation
        match = has_person
        person_part = match.group(1).strip()
        
        # Ensure person part ends with period
        if not person_part.endswith('.'):
            person_part += '.'
        
        # Get the rest (product description)
        rest = alt_text.replace(match.group(1), '').strip()
        
        # Remove leading punctuation from rest if any
        rest = re.sub(r'^[,.\s]+', '', rest)
        
        # Ensure rest starts with capital letter
        if rest and rest[0].islower():
            rest = rest[0].upper() + rest[1:]
        
        # Ensure rest is a complete sentence
        if rest and not rest.endswith('.'):
            rest += '.'
        
        # Combine with proper separation
        if rest:
            return f"{person_part} {rest}"
        else:
            return person_part
    
    return alt_text

# Function to clean and validate alt text
def clean_and_validate_alt_text(alt_text, device_name=""):
    """Clean, validate, and enforce all rules on alt text"""
    # Remove markdown formatting
    alt_text = re.sub(r'[*_`]', '', alt_text)
    
    # Remove quotes if wrapped
    if alt_text.startswith('"') and alt_text.endswith('"'):
        alt_text = alt_text[1:-1]
    
    # Remove brand repetition
    alt_text = remove_brand_repetition(alt_text)
    
    # Combine person holding device OR separate properly
    alt_text = separate_person_description(alt_text)
    
    # Ensure proper article/number word prefix
    alt_text = ensure_article_prefix(alt_text)
    
    # Apply device name substitution if provided
    alt_text = substitute_device_name(alt_text, device_name)
    
    # Remove any remaining brand repetition after substitution
    alt_text = remove_brand_repetition(alt_text)
    
    # Enforce 150 character limit
    alt_text = enforce_character_limit(alt_text, 150)
    
    # Final cleanup of extra spaces
    alt_text = re.sub(r'\s+', ' ', alt_text).strip()
    
    return alt_text

# Function to generate alt text for single image
def generate_alt_text(client, image_base64, filename, device_name="", model=None):
    """Generate alt text using Claude Vision API"""
    if model is None:
        model = st.session_state.selected_model
    
    try:
        start_time = time.time()
        
        media_type = get_image_media_type(filename)
        
        max_tokens = 512
        
        response = client.messages.create(
            model=model,
            max_tokens=max_tokens,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": media_type,
                                "data": image_base64,
                            },
                        },
                        {
                            "type": "text",
                            "text": ALT_TEXT_PROMPT
                        }
                    ],
                }
            ],
        )
        
        end_time = time.time()
        execution_time = end_time - start_time
        
        # Calculate cost
        cost_info = calculate_cost(response.usage, model)
        
        # Get generated alt text and clean it
        alt_text = response.content[0].text.strip()
        
        # Clean and validate
        alt_text = clean_and_validate_alt_text(alt_text, device_name)
        
        # Check for banned words in final output
        banned_found = check_banned_words(alt_text)
        if banned_found:
            # Remove banned words and regenerate
            for word in banned_found:
                pattern = r'\b' + re.escape(word) + r'\b'
                alt_text = re.sub(pattern, '', alt_text, flags=re.IGNORECASE)
            alt_text = re.sub(r'\s+', ' ', alt_text).strip()
            alt_text = clean_and_validate_alt_text(alt_text, device_name)
        
        # Update session state
        st.session_state.api_calls += 1
        st.session_state.total_cost += cost_info['total_cost']
        st.session_state.total_tokens['input'] += cost_info['input_tokens']
        st.session_state.total_tokens['output'] += cost_info['output_tokens']
        st.session_state.total_execution_time += execution_time
        
        return {
            "alt_text": alt_text,
            "cost": cost_info,
            "execution_time": execution_time
        }
    
    except Exception as e:
        return {
            "alt_text": f"Error: {str(e)}",
            "cost": None,
            "execution_time": 0
        }

# ============================================================================
# IMPROVED: Smart Batch Processing with Automatic Batching
# ============================================================================
def generate_bulk_alt_text(client, images_data, device_name="", model=None):
    """
    Generate alt text for multiple images with automatic batching.
    Splits large batches into smaller chunks to avoid API limits.
    """
    if model is None:
        model = st.session_state.selected_model
    
    # Split images into batches
    total_images = len(images_data)
    num_batches = (total_images + MAX_IMAGES_PER_BATCH - 1) // MAX_IMAGES_PER_BATCH
    
    all_alt_texts = []
    total_execution_time = 0
    total_cost_info = {
        'input_tokens': 0,
        'output_tokens': 0,
        'total_tokens': 0,
        'input_cost': 0.0,
        'output_cost': 0.0,
        'total_cost': 0.0
    }
    
    # Process each batch
    for batch_idx in range(num_batches):
        start_idx = batch_idx * MAX_IMAGES_PER_BATCH
        end_idx = min((batch_idx + 1) * MAX_IMAGES_PER_BATCH, total_images)
        
        batch_images = images_data[start_idx:end_idx]
        
        try:
            start_time = time.time()
            
            content = []
            
            # Add images from this batch
            for img_base64, filename in batch_images:
                media_type = get_image_media_type(filename)
                content.append({
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": img_base64,
                    },
                })
            
            # Enhanced batch prompt
            batch_prompt = f"""You are an expert in visual accessibility and e-commerce product description. Generate accurate, vivid, and highly detailed alt-text for each of the {len(batch_images)} product images above.

**PRODUCT IDENTIFICATION REFERENCE:**

{PMI_PRODUCT_GUIDE}

**CRITICAL REQUIREMENTS:**
1. Single product: Start with "A" or "An"
2. Multiple products (2+): Start with NUMBER WORD (Two, Three...), NOT "A" or "An"
3. Person holding/using device: COMBINE in ONE sentence with "holding" or "using"
   • Format: "A person wearing [clothing] holding a [device]"
   • NEVER separate like: "A person [desc]. A device in hand."
4. Person NOT holding: Use separate sentences
5. Include nuanced details: textures, finishes, materials, specific colors
6. **For PMI products**: Use EXACT spelling (IQOS, ILUMA, TEREA, LEVIA, VEEV, ZYN, etc.)
7. **For PMI products**: Include device model and compatibility when identifiable
8. NO marketing language or subjective claims
9. **MAXIMUM 150 characters per alt-text**
10. **NEVER repeat brand names** (e.g., avoid "VEEV VEEV")
11. **STRICTLY AVOID BANNED WORDS**: love, adore, relish, don't miss, luxurious, best, better, worse, worst, safe, safer, safest, fastest, fresh, seamless, lovely, adorable, sophisticated, beautiful, satisfying, cheap, affordable, low cost, lower price, lowest prices, budget-friendly, economical, inexpensive

**Output Format:**
Return a numbered list (1., 2., 3., etc.) with one alt-text per line.
Each line: ONLY the alt-text string (max 150 characters, no quotes).

Examples:
1. A person wearing a yellow shirt holding a black and silver device with metallic finish
2. Two stainless steel devices in silver finish with LED indicators
3. IQOS ILUMA PRIME holder in moonlight silver with pocket charger, bladeless heated tobacco device
4. TEREA Amber tobacco sticks pack in orange packaging for IQOS ILUMA devices"""

            content.append({
                "type": "text",
                "text": batch_prompt
            })
            
            # Adjust max_tokens for this batch
            max_tokens = 512 * len(batch_images)
            
            # API call for this batch
            response = client.messages.create(
                model=model,
                max_tokens=max_tokens,
                messages=[
                    {
                        "role": "user",
                        "content": content
                    }
                ],
            )
            
            end_time = time.time()
            execution_time = end_time - start_time
            total_execution_time += execution_time
            
            # Calculate cost for this batch
            cost_info = calculate_cost(response.usage, model)
            
            # Accumulate total costs
            total_cost_info['input_tokens'] += cost_info['input_tokens']
            total_cost_info['output_tokens'] += cost_info['output_tokens']
            total_cost_info['total_tokens'] += cost_info['total_tokens']
            total_cost_info['input_cost'] += cost_info['input_cost']
            total_cost_info['output_cost'] += cost_info['output_cost']
            total_cost_info['total_cost'] += cost_info['total_cost']
            
            # Parse response for this batch
            response_text = response.content[0].text
            batch_alt_texts = []
            
            lines = response_text.strip().split('\n')
            for line in lines:
                match = re.match(r'^\d+[\.\)]\s*(.+)$', line.strip())
                if match:
                    alt_text = match.group(1).strip()
                    alt_text = clean_and_validate_alt_text(alt_text, device_name)
                    
                    # Check for banned words
                    banned_found = check_banned_words(alt_text)
                    if banned_found:
                        for word in banned_found:
                            pattern = r'\b' + re.escape(word) + r'\b'
                            alt_text = re.sub(pattern, '', alt_text, flags=re.IGNORECASE)
                        alt_text = re.sub(r'\s+', ' ', alt_text).strip()
                        alt_text = clean_and_validate_alt_text(alt_text, device_name)
                    
                    batch_alt_texts.append(alt_text)
            
            # Fallback parsing
            if len(batch_alt_texts) != len(batch_images):
                batch_alt_texts = [line.strip() for line in lines if line.strip() and not line.strip().startswith('#')]
                batch_alt_texts = [clean_and_validate_alt_text(text, device_name) for text in batch_alt_texts]
            
            # Pad or truncate
            if len(batch_alt_texts) < len(batch_images):
                batch_alt_texts.extend([f"A {batch_images[i][1]}"[:150] for i in range(len(batch_alt_texts), len(batch_images))])
            elif len(batch_alt_texts) > len(batch_images):
                batch_alt_texts = batch_alt_texts[:len(batch_images)]
            
            # Add to overall results
            all_alt_texts.extend(batch_alt_texts)
            
            # Update session state for this batch
            st.session_state.api_calls += 1
            st.session_state.total_cost += cost_info['total_cost']
            st.session_state.total_tokens['input'] += cost_info['input_tokens']
            st.session_state.total_tokens['output'] += cost_info['output_tokens']
            st.session_state.total_execution_time += execution_time
            
        except Exception as e:
            # If batch fails, add error messages for all images in batch
            error_texts = [f"Error: {str(e)}"[:150]] * len(batch_images)
            all_alt_texts.extend(error_texts)
    
    return {
        "alt_texts": all_alt_texts,
        "cost": total_cost_info,
        "execution_time": total_execution_time,
        "success": True,
        "num_api_calls": num_batches
    }

# Main app logic
if api_key:
    client = Anthropic(api_key=api_key)
    
    mode = st.radio(
        "Select Mode:",
        ["Single Image", "Bulk Upload"],
        horizontal=True
    )
    
    with st.expander("🔖 Override Device Name (Optional)"):
        device_name_input = st.text_input(
            "Device name is:",
            placeholder="E.g., VEEV ONE, IQOS ILUMA, TEREA",
            help="Replace detected device names with this specific name. For PMI products, use exact spelling (IQOS, ILUMA, TEREA, LEVIA, VEEV, ZYN, etc.)",
            key="device_name_input"
        )
        
        # Check for banned words in device name
        if device_name_input:
            banned_in_device = check_banned_words(device_name_input)
            if banned_in_device:
                st.error(f"❌ **Banned words detected:** {', '.join(banned_in_device)}")
                st.warning("Please remove banned words from the device name.")
                device_name = ""
            else:
                device_name = device_name_input
                st.info(f"✅ Device name will be applied to all products in images (with automatic pluralization)")
        else:
            device_name = ""
    
    # Display banned words reference
    with st.expander("🚫 Banned Words Reference"):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("**Banned Verbs:**")
            for word in BANNED_VERBS:
                st.caption(f"• {word}")
        
        with col2:
            st.markdown("**Banned Adjectives:**")
            for word in BANNED_ADJECTIVES:
                st.caption(f"• {word}")
        
        with col3:
            st.markdown("**Banned Price Terms:**")
            for word in BANNED_PRICE_TERMS:
                st.caption(f"• {word}")
    
    # Display PMI product reference
    with st.expander("📚 PMI Product Reference Guide"):
        st.markdown(PMI_PRODUCT_GUIDE)
    
    st.markdown("---")
    
    if mode == "Single Image":
        st.subheader("Upload Single Image")
        
        uploaded_file = st.file_uploader(
            "Choose an image",
            type=["jpg", "jpeg", "png", "webp"],
            key="single"
        )
        
        if uploaded_file:
            col1, col2 = st.columns(2)
            
            with col1:
                st.image(uploaded_file, caption="Uploaded Image", use_container_width=True)
            
            with col2:
                if st.button("Generate Alt Text", type="primary", use_container_width=True):
                    # Check if device name has banned words
                    if device_name and check_banned_words(device_name):
                        st.error("❌ Cannot generate: Device name contains banned words. Please remove them first.")
                    else:
                        with st.spinner(f"Generating with {MODELS[st.session_state.selected_model]}..."):
                            uploaded_file.seek(0)
                            image_base64 = encode_image(uploaded_file)
                            result = generate_alt_text(client, image_base64, uploaded_file.name, device_name)
                            
                            # Check for banned words in result
                            banned_in_result = check_banned_words(result['alt_text'])
                            if banned_in_result:
                                st.warning(f"⚠️ Banned words detected in output: {', '.join(banned_in_result)}")
                                st.info("Attempting to clean...")
                            
                            st.session_state.current_result = result
                            st.rerun()
            
            if st.session_state.current_result:
                result = st.session_state.current_result
                
                st.markdown("---")
                st.subheader("📊 Session Statistics")
                stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
                
                with stat_col1:
                    st.metric("API Calls", st.session_state.api_calls)
                with stat_col2:
                    st.metric("⏱️ Total Time", f"{st.session_state.total_execution_time:.2f}s")
                with stat_col3:
                    st.metric("💰 Total Cost", f"${st.session_state.total_cost:.4f}")
                with stat_col4:
                    if st.session_state.api_calls > 0:
                        avg_cost = st.session_state.total_cost / st.session_state.api_calls
                        st.metric("Avg Cost", f"${avg_cost:.4f}")
                
                st.markdown("---")
                
                # Check for banned words in result
                banned_in_output = check_banned_words(result['alt_text'])
                if banned_in_output:
                    st.error(f"❌ **Warning: Banned words found:** {', '.join(banned_in_output)}")
                else:
                    st.success("✅ Alt text generated (no banned words detected)!")
                
                st.text_area(
                    "Generated Alt Text:",
                    value=result['alt_text'],
                    height=150,
                    key="result_display"
                )
                
                char_count = len(result['alt_text'])
                if char_count > 150:
                    st.error(f"❌ Character count: {char_count} (exceeds 150 limit)")
                else:
                    st.success(f"✅ Character count: {char_count} / 150")
                
                if result['cost']:
                    st.markdown("---")
                    st.subheader("This Request")
                    col_a, col_b, col_c = st.columns(3)
                    
                    with col_a:
                        st.metric("⏱️ Execution Time", f"{result['execution_time']:.2f}s")
                    with col_b:
                        st.metric("💰 Cost", f"${result['cost']['total_cost']:.6f}")
                    with col_c:
                        st.metric("🔢 Total Tokens", f"{result['cost']['total_tokens']:,}")
                    
                    with st.expander("📊 Detailed Token Usage"):
                        st.write(f"**Input tokens:** {result['cost']['input_tokens']:,} (${result['cost']['input_cost']:.6f})")
                        st.write(f"**Output tokens:** {result['cost']['output_tokens']:,} (${result['cost']['output_cost']:.6f})")
    
    else:  # Bulk Upload mode
        st.subheader("Upload Multiple Images")
        
        uploaded_files = st.file_uploader(
            "Choose images",
            type=["jpg", "jpeg", "png", "webp"],
            accept_multiple_files=True,
            key="bulk"
        )
        
        if uploaded_files:
            st.info(f"📁 {len(uploaded_files)} image(s) uploaded")
            
            # Show optimization notice
            num_batches = (len(uploaded_files) + MAX_IMAGES_PER_BATCH - 1) // MAX_IMAGES_PER_BATCH
            
            if len(uploaded_files) > 1:
                if num_batches > 1:
                    st.success(f"✨ Smart batching enabled: {len(uploaded_files)} images will be processed in {num_batches} batches ({MAX_IMAGES_PER_BATCH} per call)")
                else:
                    st.success(f"✨ Batch optimization enabled: All {len(uploaded_files)} images will be processed in 1 API call")
            
            if st.button("Generate Alt Text for All", type="primary"):
                # Check if device name has banned words
                if device_name and check_banned_words(device_name):
                    st.error("❌ Cannot generate: Device name contains banned words. Please remove them first.")
                else:
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    batch_start_time = time.time()
                    
                    status_text.text("Encoding images...")
                    images_data = []
                    
                    for file in uploaded_files:
                        file.seek(0)
                        images_data.append((encode_image(file), file.name))
                    
                    progress_bar.progress(0.3)
                    
                    if num_batches > 1:
                        status_text.text(f"Processing {len(uploaded_files)} images in {num_batches} batches ({MAX_IMAGES_PER_BATCH} images per API call)...")
                    else:
                        status_text.text(f"Processing {len(uploaded_files)} images in 1 API call...")
                    
                    # Smart batch processing
                    bulk_result = generate_bulk_alt_text(client, images_data, device_name)
                    
                    progress_bar.progress(0.9)
                    
                    results = []
                    if bulk_result['success']:
                        for idx, (img_data, alt_text) in enumerate(zip(images_data, bulk_result['alt_texts'])):
                            filename = img_data[1]
                            per_image_cost = bulk_result['cost']['total_cost'] / len(images_data) if bulk_result['cost'] and bulk_result['cost']['total_cost'] > 0 else 0.0
                            per_image_tokens = bulk_result['cost']['total_tokens'] // len(images_data) if bulk_result['cost'] else 0
                            
                            # Check for banned words
                            banned_found = check_banned_words(alt_text)
                            
                            results.append({
                                "Filename": filename,
                                "Alt Text": alt_text,
                                "Character Count": len(alt_text),
                                "Est. Cost ($)": f"{per_image_cost:.6f}",
                                "Est. Tokens": per_image_tokens,
                                "Banned Words": ', '.join(banned_found) if banned_found else "None"
                            })
                    
                    batch_end_time = time.time()
                    total_batch_time = batch_end_time - batch_start_time
                    batch_cost = bulk_result['cost']['total_cost'] if bulk_result['cost'] else 0.0
                    
                    progress_bar.progress(1.0)
                    
                    # Show API call optimization
                    if bulk_result.get('num_api_calls', 1) > 1:
                        status_text.text(f"✅ Processed {len(uploaded_files)} images in {bulk_result['num_api_calls']} API calls!")
                    elif len(uploaded_files) > 1:
                        status_text.text(f"✅ Processed {len(uploaded_files)} images in 1 API call!")
                    else:
                        status_text.text("✅ All images processed!")
                    
                    st.session_state.bulk_results = {
                        'results': results,
                        'total_batch_time': total_batch_time,
                        'batch_cost': batch_cost,
                        'uploaded_files': uploaded_files,
                        'is_optimized': bulk_result['success'],
                        'num_api_calls': bulk_result.get('num_api_calls', 0)
                    }
                    st.rerun()
        
        if st.session_state.bulk_results:
            bulk_data = st.session_state.bulk_results
            results = bulk_data['results']
            total_batch_time = bulk_data['total_batch_time']
            batch_cost = bulk_data['batch_cost']
            is_optimized = bulk_data.get('is_optimized', False)
            uploaded_files_list = bulk_data.get('uploaded_files', [])
            num_api_calls = bulk_data.get('num_api_calls', 0)
            
            # Show optimization status
            if is_optimized and num_api_calls > 1:
                st.success(f"✨ Smart batching: {len(results)} images processed in {num_api_calls} API calls!")
            elif is_optimized and len(results) > 1:
                st.success(f"✨ Optimized batch processing: {len(results)} images processed in 1 API call!")
            elif is_optimized:
                st.success("✨ Optimized batch processing used!")
            
            # Check for any banned words in results
            any_banned = any(result['Banned Words'] != "None" for result in results)
            if any_banned:
                st.error("⚠️ **Warning:** Some results contain banned words. Please review below.")
            
            st.markdown("---")
            st.subheader("📊 Session Statistics")
            stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
            
            with stat_col1:
                st.metric("API Calls", st.session_state.api_calls)
            with stat_col2:
                st.metric("⏱️ Total Time", f"{st.session_state.total_execution_time:.2f}s")
            with stat_col3:
                st.metric("💰 Total Cost", f"${st.session_state.total_cost:.4f}")
            with stat_col4:
                if st.session_state.api_calls > 0:
                    avg_cost = st.session_state.total_cost / st.session_state.api_calls
                    st.metric("Avg Cost", f"${avg_cost:.4f}")
            
            st.markdown("---")
            st.subheader("📊 This Batch")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Images Processed", len(results))
            with col2:
                st.metric("Total Time", f"{total_batch_time:.2f}s")
            with col3:
                # FIX: Prevent division by zero error
                if len(results) > 0:
                    st.metric("Avg Time/Image", f"{total_batch_time/len(results):.2f}s")
                else:
                    st.metric("Avg Time/Image", "0.00s")
            with col4:
                st.metric("Batch Cost", f"${batch_cost:.4f}")
            
            st.markdown("---")
            st.subheader("Results")
            
            # Display table with thumbnails
            for idx, (result, file) in enumerate(zip(results, uploaded_files_list)):
                with st.container():
                    col1, col2, col3, col4, col5, col6 = st.columns([1, 2, 6, 1, 1, 2])
                    
                    with col1:
                        file.seek(0)
                        thumbnail = create_thumbnail(file, size=(80, 80))
                        st.image(thumbnail, use_container_width=True)
                    
                    with col2:
                        st.markdown(f"**{result['Filename']}**")
                    
                    with col3:
                        # Highlight if has banned words
                        if result['Banned Words'] != "None":
                            st.markdown(f"⚠️ {result['Alt Text']}")
                        else:
                            st.markdown(f"{result['Alt Text']}")
                    
                    with col4:
                        char_count = result['Character Count']
                        if char_count > 150:
                            st.markdown(f"❌ {char_count}")
                        else:
                            st.markdown(f"✅ {char_count}")
                    
                    with col5:
                        st.markdown(f"💰 {result['Est. Cost ($)']}")
                    
                    with col6:
                        if result['Banned Words'] != "None":
                            st.markdown(f"🚫 {result['Banned Words']}")
                        else:
                            st.markdown("✅ Clean")
                    
                    st.divider()
            
            # Download buttons
            st.markdown("---")
            col_csv, col_excel = st.columns(2)
            
            with col_csv:
                # CSV download
                df = pd.DataFrame(results)
                csv = df.to_csv(index=False)
                st.download_button(
                    label="📥 Download as CSV",
                    data=csv,
                    file_name=f"alt_text_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            with col_excel:
                # Excel download (with thumbnails)
                excel_file = create_excel_with_images(results, uploaded_files_list)
                st.download_button(
                    label="📊 Download as Excel (with images)",
                    data=excel_file,
                    file_name=f"alt_text_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            st.markdown("---")
            st.subheader("Preview")
            
            cols = st.columns(3)
            for idx, (file, result) in enumerate(zip(uploaded_files_list, results)):
                with cols[idx % 3]:
                    file.seek(0)
                    st.image(file, caption=file.name, use_container_width=True)
                    
                    # Show warning if banned words or over limit
                    if result['Banned Words'] != "None":
                        st.error(f"🚫 Banned: {result['Banned Words']}")
                    if result['Character Count'] > 150:
                        st.error(f"❌ Length: {result['Character Count']}")
                    
                    st.caption(f"**Alt:** {result['Alt Text']}")
                    st.caption(f"💰 {result['Est. Cost ($)']}")
                    st.markdown("---")

else:
    st.info("👈 Please enter your Anthropic API key in the sidebar to get started")
    
    st.markdown("### How to use:")
    st.markdown("""
    1. Get your API key from [console.anthropic.com](https://console.anthropic.com/)
    2. Enter it in the sidebar
    3. Select model:
       - **Opus 4** (default): Premium quality, top results
       - **Haiku 3.5**: Fast and cost-effective alternative
    4. Adjust batch size (3-10 images per API call)
    5. Upload images
    6. Generate alt text
    7. Download results as CSV or Excel (with thumbnails)
    
    **✨ Key Features:**
    - **Smart Batch Processing** - Automatic batching for unlimited images
    - **Configurable batch size** - Adjust from 3-10 images per call
    - **PMI Product Identification** - IQOS, ILUMA, TEREA, LEVIA, VEEV, ZYN
    - **Multiple Devices** - Start with number word: "Two devices"
    - **Person Holding** - Combined format: "A person wearing [clothing] holding a [device]"
    - **150 character limit** - Concise, focused descriptions
    - **Banned word filtering** - Removes marketing language
    - **No brand repetition** - "VEEV" not "VEEV VEEV"
    - **Auto-pluralization** - Device name override applies to all products
    - **Nuanced details** - Textures, finishes, materials, colors
    - **Excel with thumbnails** - Visual results in spreadsheet
    - **Real-time cost tracking**
    """)

st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: gray;'>Alt-Text Wizard 🧙‍♂️✨ – Powered by Anthropic Claude</div>",
    unsafe_allow_html=True
)